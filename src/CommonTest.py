import openpyxl
import pandas as pd
from enum import IntEnum, Enum

class _Column(IntEnum):
    ROW     = 1
    NAME    = 2
    EXP_VAL = 3
    OPER    = 4
    ACT_VAL = 5
    CHECK   = 6

class _Result(Enum):
    PASS = "PASS"
    FAIL = "FAIL"

class _Operation(Enum):
    EQ   = ("="  , "Equals")
    SEQ  = ("="  , "Equals")
    NE   = ("<>" , "Not Equals")
    NEQ  = ("<>" , "Not Equals")
    GE   = (">=" , "Greater Than or Equals")
    GT   = (">"  , "Greater Than")
    LE   = ("<=" , "Less Than or Equals")
    LT   = ("<"  , "Less Than")
    TL   = ("+/-", "Within")
    NTL  = ("+/-", "Not Within")

class _CellFormat(Enum):
    REDFONT = openpyxl.styles.Font(
        color="0000FF", 
        bold=True)
    
    ORANGEFILL = openpyxl.styles.PatternFill(
        start_color="FFA500", 
        end_color="FFA500",
        fill_type="solid")
    
    HYPERLINK = openpyxl.styles.Font(
        color="0000FF", 
        underline="single")

class CommonTest:
    def __init__(self):
        # Workbook / worksheet state
        self.workbook   = None
        self.results_ws = None
        self.active_ws  = None
        self.currentResultsRow = None
        self.total_rows = []
        self.total_cols = []

        # Default row/column start (optional)
        self._rowCol    = _Column.ROW
        self._nameCol   = _Column.NAME
        self._expValCol = _Column.EXP_VAL
        self._operCol   = _Column.OPER
        self._actValCol = _Column.ACT_VAL
        self._checkCol  = _Column.CHECK

        # Default result values
        self._passVal = _Result.PASS
        self._failVal = _Result.FAIL

        # _Operations dictionary for quick lookup
        self._basicOperations = {op.name: op.value for op in _Operation}

        # Color Formats
        self._redFont = _CellFormat.REDFONT
        self._orangeFill = _CellFormat.ORANGEFILL
        self._hyperLinkFont = _CellFormat.HYPERLINK   

    def initializeTest(self, csv_files, output_file):
        with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
            self.workbook = writer.book
            header_format = self.workbook.add_format({'bold': True})
            for index, csv_file in enumerate(csv_files):
                sheet_name = f"AnalyzedData-{index+1}"
                df = pd.read_csv(csv_file, dtype=str)
                df.to_excel(writer, sheet_name=sheet_name, index=False)   
                self.active_ws = writer.sheets[sheet_name]

                for col_num, value in enumerate(df.columns.values):
                    self.active_ws.write(0, col_num, value, header_format)
                
                for i, col_num  in enumerate(df.columns):
                    max_length = max(df[col_num].astype(str).str.len().max(), len(col_num))

                    self.active_ws.set_column(i, i, max_length + 2)

                self.active_ws.freeze_panes(1,0)
                rows, cols = df.shape
                
                self.total_rows.append(rows)
                self.total_cols.append(cols)
        
            self._createResultsFile()
        self._openResultsWorkbook(output_file)

    def endTest(self, output_file):
        self.workbook.save(output_file) 

    def _createResultsFile(self):
        self.results_ws = self.workbook.add_worksheet("Results")
        self._formatResultsWorksheet()

    def _formatResultsWorksheet(self):
        header_format = self.workbook.add_format({
            "bold": True,
            "align": "center",
            "valign": "vcenter",
            "border": 1
        })

        headers = ["Row Index","Name","Actual Value",
                    "Operation","Expected Value"]

        for col_num, header in enumerate(headers):
            self.results_ws.write(0, col_num, header, header_format)

        self.results_ws.freeze_panes(1,0)

    def _openResultsWorkbook(self, output_file):
        self.workbook   = openpyxl.load_workbook(output_file)
        self.active_ws  = self.workbook.active
        self.results_ws = self.workbook['Results']       

    def _activateWorksheet(self, csvSheet):
        self.active_ws = self.workbook[f"AnalyzedData-{csvSheet}"]
    
    def getRowNumber(self, searchString: str, colNum, csvSheet=1, startRow=None):
        self._activateWorksheet(csvSheet)
        startRow = startRow or self.active_ws.min_row
        iterator = self.active_ws.iter_rows(min_col=colNum, 
                                            max_col=colNum,
                                            min_row=startRow,
                                            values_only=True)
        
        for rowNum, rowVal in enumerate(iterator, start=startRow):
            if rowVal[0] == searchString:
                return rowNum
    
    def getColNumber(self, searchString, csvSheet=1):
        self._activateWorksheet(csvSheet)
        iterator = self.active_ws.iter_rows(min_row=1, 
                                            max_row=1,
                                            values_only=True)
        
        for colNum, colVal in enumerate(iterator, start=self.active_ws.min_column):
            if colVal[0] == searchString:
                return colNum

    def findAllRows(self, searchString, colNum, csvSheet=1):   
        allRows = []
        self._activateWorksheet(csvSheet)
        iterator = self.active_ws.iter_rows(min_col=colNum, 
                                            max_col=colNum,
                                            values_only=True)
        
        for rowNum, rowVal in enumerate(iterator, start=self.active_ws.min_row):
            if rowVal[0] == searchString:
                allRows.append(rowNum)
        
        return allRows

    def findRowsIntersect(self, searchStringDict, csvSheet=1):
        intersection = None
        for key, value in searchStringDict.items():
            allRows = set(self.findAllRows(searchString=value, 
                                           colNum=key, 
                                           csvSheet=csvSheet))
            
            intersection =  allRows if intersection is None else (intersection & allRows)

        return list(intersection)

    def findRowsUnion(self, searchStringDict, csvSheet=1):
        union = set()
        for key, value in searchStringDict.items():
            union |= set(self.findAllRows(searchString=value,
                                          colNum=key, 
                                          csvSheet=csvSheet))

        return list(union)
    
    def _returnStrCoordRC(self):
        actVal = f'RC[{self._actValCol - self._checkCol}]'
        expVal = f'RC[{self._expValCol - self._checkCol}]'
        return actVal, expVal
    
    def _basicFormula(self, operator):
        actVal, expVal = self._returnStrCoordRC()
        return f"=IF({actVal} {operator} {expVal}, {self._passVal}, {self._failVal})"

    def _commandTolerance(self, tol, cmmd):
        actVal, expVal = self._returnStrCoordRC()
        check1 = f"{actVal} <= {expVal} + {tol}"
        check2 = f"{actVal} => {expVal} - {tol}"
        formula = f"AND({check1}, {check2})"
        
        if cmmd in ("NTL"):
            formula = f"NOT({formula})"

        return f"=IF({formula}, {self._passVal}, {self._failVal})"

    def expectedValuesCheck(self, expectedValue, actualValue):
        expVal = None
        stringSplit = expectedValue.trim().split(",", 1)
        commandStr = stringSplit[0]
        symbol, description = self._basicOperations.get(commandStr,
                                                        (None, None))
        if commandStr in ("TL","NTL"):
            parts = stringSplit[1].rsplit(",", 1)
            expVal, tolerance = parts[0], parts[1]
            val = self._commandWithtinTolerance(tol=tolerance, cmmd=commandStr)
            description = f"{description} {symbol} {tolerance}"
        
        elif commandStr in self._basicOperations:
            val = self._basicFormula(operator=symbol)
            expVal = stringSplit[1]
        
        else:
            val = None
        
        # Writing Information to Results Worksheet
        self._setResultCellValue(value=expVal,
                                 resultRow=self.currentResultsRow,
                                 resultCol=self._expValCol)
        
        self._setResultCellValue(value=actualValue,
                                 resultRow=self.currentResultsRow,
                                 resultCol=self._actValCol)
        
        self._setResultCellValue(value=description,
                                 resultRow=self.currentResultsRow,
                                 resultCol=self._operCol)
        
        self._setResultCellValue(value=val,
                                 resultRow=self.currentResultsRow,
                                 resultCol=self._checkCol)

    def _increaseResultsRow(self):   
        self.currentResultsRow += 1

    def setCellValue(self, row, col, value, csvSheet=1):
        self._activateWorksheet(csvSheet=csvSheet)
        self.active_ws(row=row, column=col, value=value).font = self._redFont

    def _setCellHyperlink(self, dataRow, dataCol, resultRow, resultCol, csvSheet):
        cell = self.results_ws(row=resultRow, column=resultCol)
        cell.hyperlink = f"#AnalyzedData-{csvSheet}!R{dataRow}C{dataCol}"
        cell.font = self._hyperLinkFont
    
    def _setResultCellValue(self, value, resultRow, resultCol):
        self.results_ws(row=resultRow, column=resultCol, value=value)

    def _setCellComment(self, value, resultRow, resultCol):
        self.results_ws(row=resultRow, column=resultCol).comment = value

    def addDataNameResults(self, titleStr, dataRow, dataCol, cmmt, csvSheet):
        self._setResultCellValue(value=titleStr,
                                       resultRow=self.currentResultsRow, 
                                       resultCol=self._nameCol)
        
        self._setResultCellValue(value=dataRow,
                                       resultRow=self.currentResultsRow, 
                                       resultCol=self._rowCol)

        self._setCellHyperlink(dataRow=dataRow, 
                               dataCol=1, 
                               resultRow=self.currentResultsRow, 
                               resultCol=self._rowCol, 
                               csvSheet=csvSheet)
        
        if dataCol > 0:
            self._setCellHyperlink(dataRow=dataRow, 
                                    dataCol=dataCol, 
                                    resultRow=self.currentResultsRow, 
                                    resultCol=self._actValCol, 
                                    csvSheet=csvSheet)
        
        if not cmmt is None:
            self._setCellComment(value=openpyxl.comments.Comment(cmmt))
            
    def writeResults(self, titleStr, dataRow, expectedValue, 
                     actualValue, dataCol=0, cmmt=None, csvSheet=1):
        
        self.addDataNameResults(titleStr=titleStr, 
                                dataRow=dataRow, 
                                dataCol=dataCol, 
                                cmmt=cmmt)
        
        self.expectedValuesCheck(expectedValue=expectedValue,
                                 actualValue=actualValue)

        self._increaseResultsRow()
        