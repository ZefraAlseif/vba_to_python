import openpyxl
import pandas as pd
from enum import IntEnum, Enum

class _Column(IntEnum):
    ROW     = 1
    NAME    = 2
    EXP_VAL = 3
    OPER    = 4
    ACT_VAL = 5
    CHECK   = 6

class _Result(Enum):
    PASS = "PASS"
    FAIL = "FAIL"

class _Operation(Enum):
    EQ  = ("=", "Equals")
    SEQ = ("=", "Equals")
    NE  = ("<>", "Not Equals")
    NEQ = ("<>", "Not Equals")
    GE  = (">=", "Greater Than or Equals")
    GT  = (">", "Greater Than")
    LE  = ("<=", "Less Than or Equals")
    LT  = ("<", "Less Than")

class _CellFormat(Enum):
    REDFONT = openpyxl.styles.Font(color="0000FF", bold=True)
    ORANGEFILL = openpyxl.styles.PatternFill(start_color="FFA500", 
                                              end_color="FFA500",
                                              ill_type="solid")

class CommonTest:
    def __init__(self):
        # Workbook / worksheet state
        self.workbook   = None
        self.results_ws = None
        self.active_ws  = None
        self.currentResultsRow = None
        self.total_rows = []
        self.total_cols = []

        # Default row/column start (optional)
        self._rowCol    = _Column.ROW
        self._nameCol   = _Column.NAME
        self._expValCol = _Column.EXP_VAL
        self._operCol   = _Column.OPER
        self._actValCol = _Column.ACT_VAL
        self._checkCol  = _Column.CHECK

        # Default result values
        self._passVal = _Result.PASS
        self._failVal = _Result.FAIL

        # _Operations dictionary for quick lookup
        self._basicOperations = {op.name: op.value for op in _Operation}

        # Color Formats
        self._redFont = _CellFormat.REDFONT
        self._orangeFill = _CellFormat.ORANGEFILL   

    def initializeTest(self, csv_files, output_file):
        with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
            self.workbook = writer.book
            header_format = self.workbook.add_format({'bold': True})
            for index, csv_file in enumerate(csv_files):
                sheet_name = f"AnalyzedData-{index}"
                df = pd.read_csv(csv_file)
                df.to_excel(writer, sheet_name=sheet_name, index=False)   
                self.active_ws = writer.sheets[sheet_name]

                for col_num, value in enumerate(df.columns.values):
                    self.active_ws.writer(0, col_num, value, header_format)
                
                for i, col_num  in enumerate(df.columns):
                    max_length = max(df[col_num].astype(str).map(len).map(), len(col_num))
                    self.active_ws.set_column(i, i, max_length + 2)

                self.active_ws.freeze_panes(1,0)
                rows, cols = df.shape
                
                self.total_rows.append(rows)
                self.total_cols.append(cols)
        
        self._createResultsFile()
        self.workbook.close()
        self._openResultsWorkbook(output_file)

    def endTest(self, output_file):
        self.workbook.save(output_file) 

    def _createResultsFile(self):
        self.results_ws = self.workbook.add("Results")
        self._formatResultsWorksheet()

    def _formatResultsWorksheet(self):
        header_format = self.workbook.add_format({
            "bold": True,
            "align": "center",
            "valign": "vcenter",
            "border": 1
        })

        headers = ["Row Index","Name","Actual Value",
                    "Operation","Expected Value"]

        for col_num, header in enumerate(headers):
            self.results_ws.write(0, col_num, header, header_format)

        self.results_ws.freeze_panes(1,0)

    def _openResultsWorkbook(self, output_file):
        self.workbook   = openpyxl.load_workbook(output_file)
        self.active_ws  = self.workbook.active
        self.results_ws = self.workbook["Results"]

    def _activateWorksheet(self, csvSheet):
        self.active_ws = self.workbook[f"AnalyzedData-{csvSheet}"]
    
    def getRowNumber(self, searchString, colNum, csvSheet):
        self._activateWorksheet(csvSheet)
        iterator = self.active_ws.iter_rows(min_col=colNum, 
                                            max_col=colNum,
                                            values_only=True)
        
        for rowNum, rowVal in enumerate(iterator, start=self.active_ws.min_row):
            if rowVal[0] == searchString:
                return rowNum

    def findAllRows(self, searchString, colNum, csvSheet):   
        allRows = []
        self._activateWorksheet(csvSheet)
        iterator = self.active_ws.iter_rows(min_col=colNum, 
                                            max_col=colNum,
                                            values_only=True)
        
        for rowNum, rowVal in enumerate(iterator, start=self.active_ws.min_row):
            if rowVal[0] == searchString:
                allRows.append(rowNum)
        
        return allRows

    def findRowsIntersect(self, searchStringDict, csvSheet):
        intersection = None
        for key, value in searchStringDict.items():
            allRows = set(self.findAllRows(searchString=value, 
                                           colNum=key, 
                                           csvSheet=csvSheet))
            
            intersection =  allRows if intersection is None else (intersection & allRows)

        return list(intersection)

    def findRowsUnion(self, searchStringDict, csvSheet):
        union = set()
        for key, value in searchStringDict.items():
            union |= set(self.findAllRows(searchString=value,
                                          colNum=key, 
                                          csvSheet=csvSheet))

        return list(union)
    
    def _returnStrCoordRC(self):
        actVal = f'RC[{self._actValCol - self._checkCol}]'
        expVal = f'RC[{self._expValCol - self._checkCol}]'
        return actVal, expVal
    
    def _basicFormula(self, operator):
        actVal, expVal = self._returnStrCoordRC()
        return f"=IF({actVal} {operator} {expVal}, {self._passVal}, {self._failVal})"

    def _commandTolerance(self, tol, cmmd):
        actVal, expVal = self._returnStrCoordRC()
        check1 = f"{actVal} <= {expVal} + {tol}"
        check2 = f"{actVal} => {expVal} - {tol}"
        formula = f"AND({check1}, {check2})"
        
        if cmmd in ("NTL"):
            formula = f"NOT({formula})"

        return f"=IF({formula}, {self._passVal}, {self._failVal})"

    def expectedValuesCheck(self, expectedValue, actualValue):
        stringSplit = expectedValue.trim().split(",", 1)
        commandStr = stringSplit[0]

        if commandStr in self._basicOperations:
            symbol, description = self._basicOperations[commandStr]
            val = self._basicFormula(operator=symbol)
        elif commandStr in ("TL","NTL"):
            lastComma = stringSplit[1].rfind(",")
            tolerance = stringSplit[1][lastComma+1:]
            val = self._commandWithtinTolerance(tol=tolerance, cmmd=commandStr)
        else:
            # unsuported command
            pass

    def _increaseResultsRow(self):   
        self.currentResultsRow += 1

    def setCellValue(self, row, col, value, csvSheet=1):
        self._activateWorksheet(csvSheet=csvSheet)
        self.active_ws(row=row, column=col, value=value).font = self._redFont

    def _setResultCellValue(self, row, col, value):
        self.results_ws(row=self.currentResultsRow, column=col, value=value)

    def addDataNameResults(self, titleStr, dataRow, colNum, cmmt):
        self._setResultCellValue(row=dataRow, col=colNum, value=titleStr)

    def writeResults(self, titleStr, dataRow, expectedValue, 
                     actualValue, colNum=None, cmmt=None):
        
        self.addDataNameResults(titleStr=titleStr, 
                                dataRow=dataRow, 
                                colNum=colNum, 
                                cmmt=cmmt)
        
        self.expectedValuesCheck(expectedValue=expectedValue,
                                 actualValue=actualValue)

        self._increaseResultsRow()
        